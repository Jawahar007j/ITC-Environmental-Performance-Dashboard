import os
import pandas as pd
from openpyxl import load_workbook

# ==================================================
# FILE PATHS
# ==================================================

DATASET = r"../DATASET/ITC_ESG_Dataset_2026.xlsx"

OUTPUT = r"../OUTPUT"

# ==================================================
# VERIFIED KPI DATA
# (Extracted from the KPI tables we validated)
# ==================================================

rows = [

# -------------------- PAGE 367 --------------------

["Packaging",
 "Waste Paper Used",
 93000,
 85000,
 "Tonnes",
 367],

["Packaging",
 "Recycled Plastic Used",
 2300,
 300,
 "Tonnes",
 367],

# -------------------- PAGE 382 --------------------

["Energy",
 "Renewable Electricity Consumption",
 964,
 868,
 "TJ",
 382],

["Energy",
 "Renewable Fuel Consumption",
 11182,
 12383,
 "TJ",
 382],

["Energy",
 "Other Renewable Energy",
 187,
 183,
 "TJ",
 382],

["Energy",
 "Total Renewable Energy",
 12334,
 13434,
 "TJ",
 382],

["Energy",
 "Non-Renewable Electricity",
 761,
 757,
 "TJ",
 382],

["Energy",
 "Non-Renewable Fuel",
 10930,
 11600,
 "TJ",
 382],

["Energy",
 "Total Non-Renewable Energy",
 11691,
 12357,
 "TJ",
 382],

["Energy",
 "Total Energy Consumed",
 24025,
 25791,
 "TJ",
 382],

# -------------------- PAGE 385 --------------------

["GHG",
 "Scope 1 Emissions",
 1045,
 1101,
 "ktCO₂e",
 385],

["GHG",
 "Scope 2 Emissions",
 151,
 153,
 "ktCO₂e",
 385],

["GHG",
 "GHG Intensity per Revenue",
 14.6,
 16.9,
 "tCO₂e/Crore INR",
 385],

["GHG",
 "GHG Intensity (PPP)",
 29.8,
 34.9,
 "tCO₂e/Million USD",
 385],

["GHG",
 "GHG Intensity (PSPD)",
 1.01,
 1.06,
 "tCO₂e/Tonne",
 385],

["GHG",
 "GHG Intensity (FBD)",
 0.16,
 0.18,
 "tCO₂e/Tonne",
 385]

]

# ==================================================
# APPEND TO EXCEL
# ==================================================

wb = load_workbook(DATASET)

ws = wb["Environmental_KPIs"]

for row in rows:

    ws.append([
        row[0],      # Category
        row[1],      # KPI
        row[2],      # FY2025-26
        row[3],      # FY2024-25
        row[4],      # Unit
        row[5],      # Source Page
        "Verified KPI Table"
    ])

wb.save(DATASET)

print("="*60)
print("Verified Environmental KPIs Added Successfully!")
print(f"Rows Added : {len(rows)}")
print("="*60)

from openpyxl import load_workbook

# ==========================================
# FILE PATH
# ==========================================

DATASET = r"../DATASET/ITC_ESG_Dataset_2026.xlsx"

# ==========================================
# VERIFIED WASTE KPIs (Page 386)
# ==========================================

rows = [

# ---------------- Waste Generated ----------------

["Waste", "Plastic Waste Generated", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "E-Waste Generated", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Biomedical Waste Generated", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Construction & Demolition Waste", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Battery Waste Generated", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Hazardous Waste Generated", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Non-Hazardous Waste Generated", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Total Waste Generated", "", "", "kilo tonnes", 386, "Verified KPI Table"],

# ---------------- Waste Recovery ----------------

["Waste", "Waste Recycled", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Waste Re-used", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Other Waste Recovery", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Total Waste Recovered", "", "", "kilo tonnes", 386, "Verified KPI Table"],

# ---------------- Waste Disposal ----------------

["Waste", "Waste Incinerated", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Waste Landfilled", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Other Waste Disposal", "", "", "kilo tonnes", 386, "Verified KPI Table"],
["Waste", "Total Waste Disposed", "", "", "kilo tonnes", 386, "Verified KPI Table"]

]

# ==========================================
# APPEND TO EXCEL
# ==========================================

wb = load_workbook(DATASET)
ws = wb["Environmental_KPIs"]

for row in rows:
    ws.append(row)

wb.save(DATASET)

print("=" * 60)
print("Waste KPIs Added Successfully!")
print(f"Rows Added : {len(rows)}")
print("=" * 60)

from openpyxl import load_workbook

# =====================================
# FILE
# =====================================

FILE = r"../FINAL_DATA/ITC_Environmental_Master.xlsx"

wb = load_workbook(FILE)
ws = wb.active

# =====================================
# PAGE 386 VERIFIED KPIs
# =====================================

rows = [

["Waste","Generated","Plastic Waste",12.60,"","kilo tonnes",386],

["Waste","Generated","E-waste",0.08,0.08,"kilo tonnes",386],

["Waste","Generated","Bio-medical Waste",0.02,0.02,"kilo tonnes",386],

["Waste","Generated","Construction & Demolition Waste",6.40,6.10,"kilo tonnes",386],

["Waste","Generated","Battery Waste",0.25,0.23,"kilo tonnes",386],

["Waste","Generated","Radioactive Waste",0,0,"kilo tonnes",386],

["Waste","Generated","Other Hazardous Waste",24.30,22.50,"kilo tonnes",386]

]

for row in rows:
    ws.append(row)

wb.save(FILE)

print("="*60)
print("Page 386 Waste KPIs Added Successfully")
print(f"Rows Added : {len(rows)}")
print("="*60)

from openpyxl import load_workbook

# =====================================
# FILE
# =====================================

FILE = r"../FINAL_DATA/ITC_Environmental_Master.xlsx"

wb = load_workbook(FILE)
ws = wb.active

# =====================================
# VERIFIED PAGE 389 KPIs
# =====================================

rows = [

[
    "GHG",
    "Scope 3",
    "Total Scope 3 Emissions",
    1442,
    1062,
    "kilo tonnes CO₂e",
    389
],

[
    "GHG",
    "Intensity",
    "Scope 3 Emissions per Rupee of Turnover",
    17.7,
    14.3,
    "tCO₂e/Crore INR",
    389
]

]

for row in rows:
    ws.append(row)

wb.save(FILE)

print("=" * 60)
print("Page 389 Scope 3 KPIs Added Successfully")
print(f"Rows Added : {len(rows)}")
print("=" * 60)

from openpyxl import load_workbook

FILE = r"../FINAL_DATA/ITC_Environmental_Master.xlsx"

wb = load_workbook(FILE)
ws = wb.active

rows = [

["ESG Ratings","CDP Water","CDP Water Rating","A List","A List","Rating",395],

["ESG Ratings","CDP Climate","CDP Climate Rating","A-","A-","Rating",395],

["ESG Ratings","CDP Forest","CDP Forest Rating","A List","","Rating",395],

["ESG Ratings","Dow Jones","Dow Jones Best-in-Class Emerging Markets Index","Included","Included","Status",395],

["ESG Ratings","MSCI","MSCI ESG Rating","AA","AA","Rating",395]

]

for row in rows:
    ws.append(row)

wb.save(FILE)

print("=" * 60)
print("Page 395 ESG Ratings Added Successfully")
print(f"Rows Added : {len(rows)}")
print("=" * 60)

